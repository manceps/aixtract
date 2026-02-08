"""Excel document converter.

Core extraction logic adapted from CAMEL-AI ExcelToolkit
(camel/toolkits/excel_toolkit.py)
Copyright 2023-2026 @ CAMEL-AI.org. All Rights Reserved.
Licensed under the Apache License, Version 2.0
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import BinaryIO, ClassVar

from aixtract.converters.base import BaseConverter
from aixtract.core.registry import ConverterRegistry
from aixtract.models.result import DocumentMetadata, ExtractionResult


@ConverterRegistry.register
class XLSXConverter(BaseConverter):
    """Extract content from Excel spreadsheets."""

    name: ClassVar[str] = "xlsx"
    supported_extensions: ClassVar[tuple[str, ...]] = (".xlsx", ".xls")
    supported_mimetypes: ClassVar[tuple[str, ...]] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    )
    requires: ClassVar[tuple[str, ...]] = ("openpyxl",)

    def extract(
        self,
        source: Path | BinaryIO | bytes,
        filename: str | None = None,
    ) -> ExtractionResult:
        """Extract content from Excel file.

        Extraction logic adapted from CAMEL ExcelToolkit.extract_excel_content().
        """
        from openpyxl import load_workbook

        content_bytes, file_path = self._read_source(source)
        wb = load_workbook(io.BytesIO(content_bytes), data_only=True)

        markdown_parts = []
        text_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            markdown_parts.append(f"## {sheet_name}\n")

            # Build table data
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])

            if rows:
                # First row as header
                headers = rows[0]
                data_rows = rows[1:]

                # Markdown table
                markdown_parts.append(
                    "| " + " | ".join(headers) + " |"
                )
                markdown_parts.append(
                    "| " + " | ".join(["---"] * len(headers)) + " |"
                )
                for row in data_rows:
                    padded = row + [""] * (len(headers) - len(row))
                    markdown_parts.append(
                        "| " + " | ".join(padded) + " |"
                    )

                text_parts.extend(
                    ["\t".join(row) for row in [headers] + data_rows]
                )

            markdown_parts.append("")

        content = "\n".join(text_parts)
        content_markdown = "\n".join(markdown_parts)

        metadata = DocumentMetadata(
            filename=filename or (file_path.name if file_path else "spreadsheet.xlsx"),
            file_path=file_path,
            format_detected="xlsx",
            converter_used=self.name,
            word_count=len(content.split()),
            char_count=len(content),
            extra={"sheet_names": wb.sheetnames, "sheet_count": len(wb.sheetnames)},
        )

        return ExtractionResult(
            success=True,
            content=content,
            content_markdown=content_markdown,
            metadata=metadata,
        )
